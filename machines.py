import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import customtkinter
from PIL import Image, ImageTk
import pandas as pd
import subprocess
import time
import os
import openpyxl

# Main window setup
root = tk.Tk()
root.title("Maschinenbelegungssteuerung")
root.geometry('1366x768')  # Adjust for 15.6-inch screen
root.attributes('-fullscreen', True)  # Enable full screen mode
root.attributes('-topmost', True)  # Always keep window on top

# Initialize the variable for the file path
file_path_var = tk.StringVar()

# Determine screen size and dynamically adjust window size
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_width = min(1366, screen_width)
window_height = min(768, screen_height)
root.geometry(f'{window_width}x{window_height}')

# Calculate percentages for the width of content and sidebar
sidebar_percentage = 0.23  # Sidebar occupies 23% of the width
sidebar_width = int(window_width * sidebar_percentage)
content_width = window_width - sidebar_width

# Style adjustments
style = ttk.Style()
style.configure("TButton", font=("Arial", 18), padding=12)
style.configure("TLabel", font=("Arial", 20), padding=12)
style.configure("TEntry", font=("Arial", 18), padding=12)

# Define style for Treeview
style.configure("Custom.Treeview", font=('Arial', 18), rowheight=49)
style.configure("Custom.Treeview.Heading", font=('Arial', 14, 'bold'))

# Define style for machine buttons
machine_button_style = 'Machine.TButton'
style.configure(machine_button_style, font=('Arial', 20, 'bold'), padding=(35, 10))

# Create an invisible button to take focus
focus_button = ttk.Button(root)
focus_button.place(x=-100, y=-100)  # Place the button outside the visible area

bemerkung_entry_height = 5  # Change this value to adjust the height

# Ensure that the `minsize` for content and sidebar are correctly set
container = ttk.Frame(root, padding="0")
container.pack(fill='both', expand=True)
container.grid_rowconfigure(0, weight=1)
container.grid_columnconfigure(0, weight=1, minsize=content_width)  # For the main content
container.grid_columnconfigure(1, weight=0, minsize=sidebar_width)  # For the sidebar

# Content area (left)
content = ttk.Frame(container, padding="0")
content.grid(row=0, column=0, sticky="nsew")
content.grid_propagate(False)  # Prevents the content area size from automatically adjusting

# Sidebar area (right) with dynamic padding
sidebar = ttk.Frame(container, padding="0")
sidebar.grid(row=0, column=1, sticky="nsew",
             padx=(0, int(window_width * 0.02)))  # Use 2% of window width for padding
sidebar.grid_propagate(False)

# Set the desired width for the sidebar
sidebar.config(width=sidebar_width)

# First create a new style specifically for sidebar buttons
sidebar_button_style = 'Sidebar.TButton'
style.configure('Sidebar.TButton', font=('Arial', 12), padding=(15, 10))

# Set the desired width for the content and sidebar areas
content_width = root.winfo_screenwidth() - sidebar_width
buttons_loaded = False

# Define machines and icons
machines = None
workbook = None
icon_paths = {
    "Erledigt": "erledigt.png",
    "Problem": "problem.png",
    "Zurücksetzen": "reset.png",
    "Excel-Datei": "mappe.png"
}

# Global variable for the last selected machine
last_selected_machine = None

# Load the icons
icons = {name: ImageTk.PhotoImage(Image.open(icon_paths[name]).resize((30, 30))) for name in icon_paths}

# Treeview for displaying data
tree = ttk.Treeview(content, style="Custom.Treeview", show='headings', selectmode="browse")
tree["columns"] = (
    "unique_id", "Artikel", "Spannung", "Stk", "Rohling", "Status", "Wkz", "Bemerkung", "Auftragsstatus")

for col in tree["columns"]:
    tree.heading(col, text=col)
    if col == "Wkz" or col == "Rohling":
        tree.column(col, width=120, anchor='center')  # Zentriere das Kreuz in der Wkz-Spalte und Rohling-Spalte
    elif col == "Auftragsstatus" or col == 'unique_id':
        tree.column(col, width=0, stretch=False)  # Verstecke die Spalte "Auftragsstatus"
    else:
        tree.column(col, width=140, anchor='w')
tree.grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=(140, 10))

# Define tags for alternating row colors
tree.tag_configure('evenrow', background='#E8E8E8')  # Light gray for even rows
tree.tag_configure('oddrow', background='white')  # White for odd rows

# Maschinen-Buttons über der Tabelle
machine_buttons_frame = customtkinter.CTkScrollableFrame(content,
                                                         orientation="horizontal",
                                                         fg_color="white",
                                                         height=80,
                                                         width=1018,
                                                         border_color="#ACACAC",
                                                         border_width=1,
                                                         corner_radius=1)  # Eckradius setzen
machine_buttons_frame.grid(row=0, column=0, sticky="nw", padx=(10, 0), pady=10)


def get_file_timestamp(file_path):
    """Returns the timestamp of the last modification of the file."""
    try:
        return os.path.getmtime(file_path)
    except OSError:
        return None


# Global variable to store the file's timestamp
last_file_timestamp = None


def check_file_change():
    global last_file_timestamp
    current_timestamp = get_file_timestamp(file_path_var.get())

    if current_timestamp is None:
        messagebox.showerror("Fehler", "Die Datei konnte nicht gefunden werden.")
        return False

    if last_file_timestamp is not None and current_timestamp != last_file_timestamp:
        load_excel_data()  # Function to load data
        last_file_timestamp = current_timestamp
        messagebox.showinfo("Aktualisierung", "Die Daten wurden aktualisiert, da die Datei extern geändert wurde.")
        return False

    return True


def try_save_excel(file_path):
    global last_file_timestamp, df, last_selected_machine
    if last_selected_machine is None:
        messagebox.showerror("No valid machine selected")
        return
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[last_selected_machine]
        # Replace 'Fräsen' with the actual name of your worksheet, if necessary

        # Disable protection
        if sheet.protection.sheet:
            sheet.protection.disable()

        # Iterate through the rows in the DataFrame and update the corresponding cells in the Excel sheet
        for index, row in df.iterrows():
            # Excel rows start at 1 and header row is added, so `index + 2`
            # Column J - Tool
            sheet.cell(row=index + 2, column=9).value = row['Wkz']
            # Assume the 9th column in the DataFrame corresponds to the tool
            # Column L - Remark
            sheet.cell(row=index + 2, column=11).value = row['Bemerkung']
            # Assume the 11th column in the DataFrame corresponds to the remark
            # Column M - Order Status
            sheet.cell(row=index + 2, column=12).value = row['Auftragsstatus']
            # Assume the 12th column in the DataFrame corresponds to the order status

        # Re-enable protection
        if sheet.protection.sheet:
            sheet.protection.enable()

        workbook.save(file_path)
        print("Workbook saved successfully")

        # Update the file's timestamp after changes have been successfully saved
        last_file_timestamp = get_file_timestamp(file_path)
        return True
    except PermissionError:
        messagebox.showerror("Fehler",
                             "Derzeit kann nicht gespeichert werden. Bitte warten Sie einen Moment und versuchen Sie "
                             "es später erneut.")
        return False
    except Exception as e:
        messagebox.showerror("Unbekannter Fehler", f"Ein unbekannter Fehler ist aufgetreten: {e}")
        return False


def on_focus_in(event):
    subprocess.Popen("osk", shell=True)


# def on_focus_out(event):
# focus_label.focus_set()

def remove_focus():
    focus_button.focus()  # Set the focus on the invisible button


# New style for the selected button
selected_machine_button_style = 'Selected.Machine.TButton'
style.configure(selected_machine_button_style, font=('Arial', 20, 'bold'), padding=(35, 10))
style.map(selected_machine_button_style,
          background=[('active', 'red'), ('!active', 'red')],
          foreground=[('active', 'red'), ('!active', 'red')])

# Global variable to store the last selected button
last_selected_button = None


def reload_machine(machine_name):
    global last_selected_machine, last_selected_button, machines

    # Check if an Excel file has been loaded
    if not file_path_var.get():
        messagebox.showerror("Fehler", "Keine Excel-Datei geladen. Bitte laden Sie zuerst eine Datei.")
        return

    if last_selected_button:
        last_selected_button.config(style=machine_button_style)

    last_selected_machine = machine_name
    global df, workbook
    data = []
    if workbook is not None:
        sheet = workbook[last_selected_machine]
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        df = pd.DataFrame(data[1:], columns=data[0])
        df.columns = df.columns.str.strip()
        df = df.drop(df[df['Artikel'] == 'Gesamtzeit'].index)

        def convert_to_int(value):
            if pd.isna(value):
                return ''
            if isinstance(value, str):
                return value
            return int(value)

        df = df.fillna('')
        # Convert 'Stückzahl' or 'Stk' to int64 if they exist and are not strings
        if 'Stückzahl' in df.columns:
            df['Stückzahl'] = df['Stückzahl'].apply(convert_to_int)

        print(f"Maschine ausgewählt: {last_selected_machine}")  # Zum Testen
        show_data(machine_name)

    # Change the style of the selected button
    for btn in machine_buttons_frame.winfo_children():
        if btn.cget("text") == machine_name:
            btn.config(style=selected_machine_button_style)
            last_selected_button = btn
            break


def select_machine(machine_name):
    global last_selected_machine, last_selected_button, machines

    # Check if an Excel file has been loaded
    if not file_path_var.get():
        messagebox.showerror("Fehler", "Keine Excel-Datei geladen. Bitte laden Sie zuerst eine Datei.")
        return

    if last_selected_button:
        last_selected_button.config(style=machine_button_style)

    last_selected_machine = machine_name
    global df, workbook
    data = []
    if workbook is not None:
        sheet = workbook[last_selected_machine]
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        df = pd.DataFrame(data[1:], columns=data[0])
        df.columns = df.columns.str.strip()
        df = df.drop(df[df['Artikel'] == 'Gesamtzeit'].index)

        def convert_to_int(value):
            if pd.isna(value):
                return ''
            if isinstance(value, str):
                return value
            return int(value)

        df = df.fillna('')
        # Convert 'Stückzahl' or 'Stk' to int64 if they exist and are not strings
        if 'Stückzahl' in df.columns:
            df['Stückzahl'] = df['Stückzahl'].apply(convert_to_int)
        if 'Stk' in df.columns:
            df['Stk'] = df['Stk'].apply(convert_to_int)

        print(f"Maschine ausgewählt: {last_selected_machine}")  # Zum Testen
        show_data(machine_name)

    # Change the style of the selected button
    for btn in machine_buttons_frame.winfo_children():
        if btn.cget("text") == machine_name:
            btn.config(style=selected_machine_button_style)
            last_selected_button = btn
            break


def ask_password():
    def on_submit():
        if password_entry.get() == "123":  # Set your password here
            focus_button.focus_set()
            select_excel_file()
            password_window.destroy()
        else:
            show_custom_warning("Error", "Falsches Passwort")
            password_window.attributes('-topmost', True)

    def show_custom_warning(title, message):
        warning_window = tk.Toplevel(password_window)
        warning_window.title(title)
        warning_window.geometry("300x150+10+10")  # Adjust size and position
        warning_window.attributes('-topmost', True)

        tk.Label(warning_window, text=message, font=("Arial", 14)).pack(padx=10, pady=10)

        ok_button = tk.Button(warning_window, text="OK", command=warning_window.destroy)
        ok_button.pack(pady=10)
        ok_button.config(font=("Arial", 14), height=2, width=10)  # Adjust button size and font

    password_window = tk.Toplevel(root)
    password_window.title("Passwort eingeben")
    password_window.geometry("400x250+10+10")  # Adjust size and position of the password window
    password_window.attributes('-topmost', True)

    password_label = ttk.Label(password_window, text="Passwort:")
    password_label.pack(padx=10, pady=10)

    password_entry = ttk.Entry(password_window, show="*")
    password_entry.pack(padx=10, pady=10)
    password_entry.bind("<FocusIn>", on_focus_in)  # Open keyboard on focus
    # password_entry.bind("<FocusOut>", on_focus_out)  # Close keyboard and remove focus

    submit_button = ttk.Button(password_window, text="Bestätigen", command=on_submit)
    submit_button.pack(padx=10, pady=10)


# Function to select the Excel file
def select_excel_file():
    global last_file_timestamp, buttons_loaded, last_selected_button
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.xlsm")])
    if file_path:
        file_path_var.set(file_path)
        load_excel_data()
        buttons_loaded = True
        last_file_timestamp = get_file_timestamp(file_path)


# Load and display data from the selected Excel file
def load_excel_data():
    global machines, workbook, buttons_loaded, last_selected_button
    if workbook:
        workbook.close()
    if file_path_var.get():
        try:
            workbook = openpyxl.load_workbook(file_path_var.get(), data_only=True)
            machines = workbook.sheetnames
            last_selected_button = None
            for widget in machine_buttons_frame.winfo_children():
                widget.destroy()
            for machine in machines:
                btn = ttk.Button(machine_buttons_frame, text=machine, style=machine_button_style,
                                 command=lambda m=machine: select_machine(m))
                btn.pack(side=tk.LEFT, padx=5, pady=5)

        except Exception as e:
            messagebox.showerror("Fehler beim Laden der Datei", str(e))


# Function to find the header and data of a specific machine

def get_machine_data(machine_name):
    headers = None
    for index, row in df.iterrows():
        headers = row[2:]  # Headers from the third column onwards
        machine_start_index = index + 1  # Data starts in the next row

    # Extract the data for the machine
    machine_data = df

    return machine_data.reset_index(drop=True)


# Function to display data for the selected machine
def show_data(machine_name):
    global last_selected_machine
    last_selected_machine = machine_name
    tree.delete(*tree.get_children())
    machine_data = None
    if machine_name:
        machine_data = get_machine_data(machine_name)
    counter = 1
    Stk_ = ''
    if 'Stk' in df.columns:
        Stk_ = 'Stk'
    else:
        Stk_ = 'Stückzahl'

    for index, row in machine_data.iterrows():
        status = row.get("Status", "")
        auftragsstatus = row.get("Auftragsstatus", "")

        # Check if the column for the item has a value
        artikel = row.get("Artikel", "")
        if artikel == "":
            continue  # Skip this row as there is no item

        # Determine tags based on the order status
        tag = ''
        if auftragsstatus == 'Erledigt':
            tag = 'erledigt'
        elif auftragsstatus == 'Problem':
            tag = 'problem'
        else:
            # Anwenden von Zebra-Streifen für Zeilen ohne speziellen Status
            tag = 'evenrow' if index % 2 == 0 else 'oddrow'

        tree.insert('', tk.END, values=(
            counter,
            artikel,
            row.get("Spannung", ""),
            row.get(Stk_, ""),
            row.get("Rohling", ""),
            status,
            row.get("Wkz", ""),
            row.get("Bemerkung", ""),
            auftragsstatus
        ), tags=(tag,))
        counter += 1

    # Configure the tags
    tree.tag_configure('evenrow', background='#E8E8E8')
    tree.tag_configure('oddrow', background='white')
    tree.tag_configure('erledigt', background='green')
    tree.tag_configure('problem', background='red')


# Function to save the remark and update the display

def save_bemerkung():
    global machines, last_selected_machine
    if machines is None:
        messagebox.showerror("Bitte wählen Sie eine Excel-Datei aus und versuchen Sie es erneut.")
        return

    if not check_file_change():
        if last_selected_machine:
            reload_machine(last_selected_machine)
        return  # Abort if the file has changed

    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Warnung", "Bitte wählen Sie einen Artikel aus.")
        return

    bemerkung_text = bemerkung_entry.get("1.0", tk.END).strip()
    if not bemerkung_text:
        messagebox.showwarning("Warnung", "Keine Bemerkung eingegeben.")
        return

    selected_item = selected_items[0]
    selected_values = tree.item(selected_item, 'values')
    unique_id = int(selected_values[0]) - 1
    selected_article_id = selected_values[1]

    for machine_name in machines:
        df_index = unique_id
        bemerkung_index = 10  # Index for "Bemerkung" in the DataFrame

        # Update remark
        df.iloc[df_index, bemerkung_index] = bemerkung_text

        # Check if the Excel file can be saved
        if not try_save_excel(file_path_var.get()):
            return  # Ends the function if the file could not be saved

        # Update the display of the row in the Treeview
        updated_values = list(tree.item(selected_item, 'values'))
        updated_values[6] = bemerkung_text  # Index of the remark in the Treeview
        tree.item(selected_item, values=updated_values)

        # Clear the remark box
        bemerkung_entry.delete("1.0", tk.END)

        print("Artikel nicht in den Daten von", machine_name, "gefunden.")
        break
    else:
        messagebox.showerror("Fehler", "Artikel wurde nicht im DataFrame gefunden.")

    reload_excel_data()


# Function to update the status and remarks
def update_status_and_bemerkung(auftragsstatus):
    global machines, last_selected_machine
    if machines is None:
        messagebox.showerror("Bitte wählen Sie eine Excel-Datei aus und versuchen Sie es erneut.")
        return

    if not check_file_change():
        if last_selected_machine:
            reload_machine(last_selected_machine)
        return  # Abort if the file has changed

    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Selection", "Please select an item.")
        return

    selected_item = selected_items[0]
    selected_values = tree.item(selected_item, 'values')
    unique_id = int(selected_values[0]) - 1
    selected_article_id = selected_values[1]

    for machine_name in machines:
        df_index = unique_id
        # Order status index in the DataFrame (18 is the index in Excel, so 17 in Python)
        auftragsstatus_df_index = 11

        if auftragsstatus == "Zurücksetzen":
            # Only reset the order status in the DataFrame to an empty string
            df.iloc[df_index, auftragsstatus_df_index] = ''
            updated_values = list(tree.item(selected_item, 'values'))
            updated_values[7] = ''  # Order status index in the Treeview list
        else:
            # Set the order status and the remark in the DataFrame
            df.iloc[df_index, auftragsstatus_df_index] = auftragsstatus
            updated_values = list(tree.item(selected_item, 'values'))
            updated_values[7] = auftragsstatus  # Order status index in the Treeview list

        # Update the display in the Treeview
        tree.item(selected_item, values=updated_values)

        # Check if the Excel file can be saved
        if not try_save_excel(file_path_var.get()):
            return  # Ends the function if the file could not be saved

        # Update the tag in the Treeview
        tag = 'erledigt' if auftragsstatus == 'Erledigt' else 'problem' if auftragsstatus == 'Problem' else ''
        tree.item(selected_item, tags=(tag,))

        print(f"Artikel {selected_article_id} nicht in den Daten von {machine_name} gefunden.")
        break
    else:
        messagebox.showerror("Fehler", "Artikel wurde nicht im DataFrame gefunden.")
    reload_excel_data()


# Function to clear the remark for the selected item
def clear_bemerkung():
    global machines, last_selected_machine
    if machines is None:
        messagebox.showerror("Bitte wählen Sie eine Excel-Datei aus und versuchen Sie es erneut.")
        return

    if not check_file_change():
        if last_selected_machine:
            reload_machine(last_selected_machine)
        return  # Abort if the file has changed

    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Warnung", "Bitte wählen Sie einen Artikel aus.")
        return

    selected_item = selected_items[0]
    selected_values = tree.item(selected_item, 'values')
    unique_id = int(selected_values[0]) - 1
    selected_article_id = selected_values[1]

    for machine_name in machines:
        df_index = unique_id

        df.iloc[df_index, 10] = ''  # Set the remark in the DataFrame to an empty string

        # Update the display of the row in the Treeview
        tree.item(selected_item, values=(tree.item(selected_item, 'values')[0:6] + ('',)))

        # Check if the Excel file can be saved
        if not try_save_excel(file_path_var.get()):
            return  # Ends the function if the file could not be saved
        messagebox.showinfo("Erfolg", "Bemerkung erfolgreich gelöscht.")
        break
    reload_excel_data()


# Adjust style for sidebar buttons
style.configure('Sidebar.TButton', font=('Arial', 12), padding=(15, 10))

# Load icons
icons = {name: ImageTk.PhotoImage(Image.open(icon_paths[name]).resize((30, 30))) for name in icon_paths}

# Create sidebar buttons in a loop for "Erledigt" and "Problem"
status_buttons_frame = ttk.Frame(sidebar)  # New frame for the status buttons
status_buttons_frame.pack(padx=0, pady=0, fill='x', expand=True)  # Place the frame in the sidebar

for status in ["Erledigt", "Problem"]:
    icon = icons[status]
    button = ttk.Button(status_buttons_frame, text=status, image=icon, style='Sidebar.TButton', compound="top",
                        command=lambda s=status: update_status_and_bemerkung(s))
    button.image = icon  # Keep reference to prevent garbage collection
    button.pack(side=tk.LEFT, padx=0, pady=0, expand=True)  # Arrange buttons side by side

# Add "Reset" button separately in the sidebar
zuruecksetzen_icon = icons["Zurücksetzen"]
zuruecksetzen_button = ttk.Button(sidebar, text="Zurücksetzen", image=zuruecksetzen_icon, style='Sidebar.TButton',
                                  compound="top", command=lambda: update_status_and_bemerkung("Zurücksetzen"))
zuruecksetzen_button.image = zuruecksetzen_icon
zuruecksetzen_button.pack(padx=0, pady=0, fill='x')  # Button below the other status buttons


def toggle_wkz():
    global machines, last_selected_machine
    if machines is None:
        messagebox.showerror("Bitte wählen Sie eine Excel-Datei aus und versuchen Sie es erneut.")
        return

    if not check_file_change():
        if last_selected_machine:
            reload_machine(last_selected_machine)
        return  # Cancel if the file has been changed

    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Auswahl", "Bitte wählen Sie einen Artikel aus.")
        return

    selected_item = selected_items[0]
    selected_values = tree.item(selected_item, 'values')
    unique_id = int(selected_values[0]) - 1
    selected_article_id = selected_values[1]

    for machine_name in machines:
        df_index = unique_id
        wkz_excel_index = 8  # Index for "Wkz" in the Excel file

        # Determine the new Wkz status
        current_wkz = df.iloc[df_index, wkz_excel_index]
        new_wkz = '' if current_wkz == 'X' else 'X'

        # Update the DataFrame, but not yet the display in the Treeview
        df.iloc[df_index, wkz_excel_index] = new_wkz

        # Attempt to save the Excel file
        if try_save_excel(file_path_var.get()):
            # Update the display of the row in the Treeview if saving was successful
            wkz_treeview_index = tree["columns"].index("Wkz")  # Index for "Wkz" in the Treeview
            updated_values = list(tree.item(selected_item, 'values'))
            updated_values[wkz_treeview_index] = new_wkz
            tree.item(selected_item, values=updated_values)

            messagebox.showinfo("Erfolg", "Wkz erfolgreich aktualisiert.")
        else:
            # If saving fails, roll back the changes in the DataFrame and display an error message
            df.iloc[df_index, wkz_excel_index] = current_wkz
            # Error message is displayed directly in try_save_excel, no need to show it again here
        break
    else:
        messagebox.showerror("Fehler", "Artikel wurde nicht im DataFrame gefunden.")
    reload_excel_data()


# Button to toggle the cross in "Wkz"
toggle_wkz_button = ttk.Button(sidebar, text="Wkz On/Off", style=sidebar_button_style, command=toggle_wkz)
toggle_wkz_button.pack(padx=0, pady=5, fill='x')

bemerkung_entry_width = 20  # Set the desired width here
bemerkung_entry_height = 9  # Set the desired height here

# Text field for remarks with on-screen keyboard support
bemerkung_entry = tk.Text(sidebar, height=bemerkung_entry_height, width=bemerkung_entry_width, font=("Arial", 18))
bemerkung_entry.bind("<FocusIn>", on_focus_in)
bemerkung_entry.bind("<FocusOut>", lambda e: remove_focus())  # Add this binding
bemerkung_entry.pack(padx=0, pady=10, fill='x')

# Buttons on the right side
# Ensure the buttons are large enough for touch interaction
button_height = 3  # Adjust the height of the buttons
button_width = 15  # Adjust the width of the buttons

# Save remark button
bemerkung_confirm_button = ttk.Button(sidebar, text="Bemerkung speichern", style=sidebar_button_style,
                                      command=save_bemerkung)
bemerkung_confirm_button.pack(padx=0, pady=5, fill='x')

# Button to clear the remark
clear_bemerkung_button = ttk.Button(sidebar, text="Bemerkung löschen", style=sidebar_button_style,
                                    command=clear_bemerkung)
clear_bemerkung_button.pack(padx=0, pady=5, fill='x')

# Label to display the last update, placed at the bottom of the window
last_updated_label = ttk.Label(root, text="Letzte Aktualisierung: -", font=("Arial", 8), borderwidth=0, relief="flat")
last_updated_label.place(relx=1.0, rely=1.0, anchor="se", height=30)


def reload_excel_data():
    global last_selected_machine
    try:
        # Save the state of the last selected machine
        saved_machine = last_selected_machine

        # Reload the data
        load_excel_data()

        # Restore the saved state
        if saved_machine:
            show_data(machine_name=saved_machine)
        else:
            show_data(None)  # Show all data if no machine was selected

        # Update the display of the last update
        current_time = time.strftime("%H:%M:%S")
        last_updated_label.config(text=f"Letzte Aktualisierung: {current_time}")
        print("Excel-Datei neu geladen um:", current_time)
    except Exception as e:
        print("Fehler beim Neuladen der Excel-Datei:", e)


def auto_reload_excel():
    # Call the function again after a set time
    root.after(300000, auto_reload_excel)  # 5-minute interval
    reload_excel_data()


# Start the automatic reload process after initializing the GUI
root.after(300000, auto_reload_excel)  # Start after 5 minutes


def open_keyboard():
    subprocess.Popen("osk", shell=True)


def ask_exit_password():
    def on_submit():
        if exit_password_entry.get() == "123":  # Set your exit password here
            root.destroy()
        else:
            messagebox.showerror("Falsches Passwort",
                                 "Das eingegebene Passwort ist falsch. Bitte versuchen Sie es erneut.",
                                 parent=exit_password_window)
            exit_password_entry.delete(0, tk.END)
            open_keyboard()  # Reopen the keyboard if the password was incorrect

    exit_password_window = tk.Toplevel()
    exit_password_window.title("Passwort zum Beenden")
    exit_password_window.geometry("300x250")  # Adjust the size
    exit_password_window.attributes('-topmost', True)

    tk.Label(exit_password_window, text="Passwort zum Beenden:").pack(padx=10, pady=10)
    exit_password_entry = ttk.Entry(exit_password_window, show="*")
    exit_password_entry.pack(padx=10, pady=10)
    exit_password_entry.bind("<FocusIn>", lambda event: open_keyboard())
    exit_password_entry.focus()  # Set focus on the input field
    submit_button = ttk.Button(exit_password_window, text="Bestätigen", command=on_submit)
    submit_button.pack(pady=10)


def on_close_request():
    ask_exit_password()


# Override the default close function of the main window
root.protocol("WM_DELETE_WINDOW", on_close_request)

# New frame for buttons in the sidebar
buttons_frame = ttk.Frame(sidebar)
buttons_frame.pack(side=tk.TOP, padx=0, pady=10)  # Pack the frame at the top of the sidebar

# Excel file button
select_file_button = ttk.Button(buttons_frame, text="Import Excel", image=icons["Excel-Datei"],
                                style=sidebar_button_style, compound="top", command=ask_password)
select_file_button.image = icons["Excel-Datei"]
select_file_button.pack(side=tk.LEFT, padx=(0, 20), pady=(0, 30))  # Place the button in the frame

# Keyboard button
keyboard_icon_image = ImageTk.PhotoImage(
    Image.open("tastatur.png").resize((30, 30)))  # Load and resize the icon
keyboard_button = ttk.Button(buttons_frame, image=keyboard_icon_image, command=open_keyboard)
keyboard_button.image = keyboard_icon_image  # Save the reference
keyboard_button.pack(side=tk.LEFT, padx=40, pady=(5, 19))  # Place the button next to the Excel file button


def ask_fullscreen_exit_password():
    def on_submit():
        if exit_password_entry.get() == "123":  # Replace with your actual password
            root.attributes("-fullscreen", False)
            exit_password_window.destroy()
        else:
            messagebox.showerror("Falsches Passwort",
                                 "Das eingegebene Passwort ist falsch. Bitte versuchen Sie es erneut.",
                                 parent=exit_password_window)
            exit_password_entry.delete(0, tk.END)

    exit_password_window = tk.Toplevel()
    exit_password_window.title("Passwort zum Beenden")
    exit_password_window.geometry("300x250")
    exit_password_window.attributes('-topmost', True)

    tk.Label(exit_password_window, text="Passwort:").pack(padx=10, pady=10)
    exit_password_entry = ttk.Entry(exit_password_window, show="*")
    exit_password_entry.pack(padx=10, pady=10)
    exit_password_entry.focus()

    submit_button = ttk.Button(exit_password_window, text="Bestätigen", command=on_submit)
    submit_button.pack(pady=10)


# Function to exit fullscreen mode
def toggle_fullscreen(event=None):
    root.attributes("-fullscreen", False)


# Bind the Escape key to exit fullscreen mode
root.bind("<Escape>", lambda event: ask_fullscreen_exit_password())

# Start the Tkinter event loop
root.mainloop()
