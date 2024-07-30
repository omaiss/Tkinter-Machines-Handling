import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import pandas as pd
import subprocess
import time
import os
import openpyxl

# Hauptfenster Setup
root = tk.Tk()
root.title("Maschinenbelegungssteuerung")
root.geometry('1366x768')  # Anpassung für 15.6 Zoll Bildschirm
root.attributes('-fullscreen', True)  # Vollbildmodus aktivieren
root.attributes('-topmost', True)  # Fenster immer im Vordergrund

# Initialisiere die Variable für den Dateipfad
file_path_var = tk.StringVar()

# Bildschirmgröße ermitteln und Fenstergröße dynamisch anpassen
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_width = min(1366, screen_width)
window_height = min(768, screen_height)
root.geometry(f'{window_width}x{window_height}')

# Berechnen Sie Prozentsätze für die Breite von Inhalten und Seitenleiste
sidebar_percentage = 0.23  # Seitenleiste nimmt 23% der Breite ein
sidebar_width = int(window_width * sidebar_percentage)
content_width = window_width - sidebar_width

# Stil-Anpassungen
style = ttk.Style()
style.configure("TButton", font=("Arial", 18), padding=12)
style.configure("TLabel", font=("Arial", 20), padding=12)
style.configure("TEntry", font=("Arial", 18), padding=12)

# Stil für Treeview definieren
style.configure("Custom.Treeview", font=('Arial', 18), rowheight=49)
style.configure("Custom.Treeview.Heading", font=('Arial', 14, 'bold'))

# Stil für die Maschinen-Buttons definieren
machine_button_style = 'Machine.TButton'
style.configure(machine_button_style, font=('Arial', 20, 'bold'), padding=(35, 10))

# Erstellen eines unsichtbaren Buttons, um den Fokus zu übernehmen
focus_button = ttk.Button(root)
focus_button.place(x=-100, y=-100)  # Platzieren Sie den Button außerhalb des sichtbaren Bereichs

bemerkung_entry_height = 5  # Ändern Sie diesen Wert, um die Höhe anzupassen

# Stellen Sie sicher, dass die `minsize` für Content und Sidebar korrekt gesetzt sind
container = ttk.Frame(root, padding="0")
container.pack(fill='both', expand=True)
container.grid_rowconfigure(0, weight=1)
container.grid_columnconfigure(0, weight=1, minsize=content_width)  # Für den Hauptinhalt
container.grid_columnconfigure(1, weight=0, minsize=sidebar_width)  # Für die Sidebar

# Content-Bereich (Links)
content = ttk.Frame(container, padding="0")
content.grid(row=0, column=0, sticky="nsew")
content.grid_propagate(False)  # Verhindert, dass die Größe des Content-Bereichs sich automatisch anpasst

# Sidebar-Bereich (Rechts) mit dynamischem Padding
sidebar = ttk.Frame(container, padding="0")
sidebar.grid(row=0, column=1, sticky="nsew",
             padx=(0, int(window_width * 0.02)))  # Verwenden Sie 2% der Fensterbreite für Padding
sidebar.grid_propagate(False)

# Setzen Sie die gewünschte Breite für die Sidebar
sidebar.config(width=sidebar_width)

# Zuerst erstellen Sie einen neuen Stil speziell für die Sidebar-Buttons
sidebar_button_style = 'Sidebar.TButton'
style.configure('Sidebar.TButton', font=('Arial', 12), padding=(15, 10))

# Setzen Sie die gewünschte Breite für den Content- und den Sidebar-Bereich
content_width = root.winfo_screenwidth() - sidebar_width

# Definiere die Maschinen und Icons
machines = ['DMU 80', 'DMU 90', 'DUO 1', 'DUO 2']
icon_paths = {
    "Erledigt": "erledigt.png",
    "Problem": "problem.png",
    "Zurücksetzen": "reset.png",
    "Excel-Datei": "mappe.png"
}

# Globale Variable für die zuletzt ausgewählte Maschine
last_selected_machine = None

# Laden der Icons
icons = {name: ImageTk.PhotoImage(Image.open(icon_paths[name]).resize((30, 30))) for name in icon_paths}

# Treeview für die Anzeige der Daten
tree = ttk.Treeview(content, style="Custom.Treeview", show='headings', selectmode="browse")
tree["columns"] = ("Artikel", "Spannung", "Stk", "Rohling", "Status", "Wkz", "Bemerkung", "Auftragsstatus")
for col in tree["columns"]:
    tree.heading(col, text=col)
    if col == "Wkz" or col == "Rohling":
        tree.column(col, width=120, anchor='center')  # Zentriere das Kreuz in der Wkz-Spalte und Rohling-Spalte
    elif col == "Auftragsstatus":
        tree.column(col, width=0, stretch=False)  # Verstecke die Spalte "Auftragsstatus"
    else:
        tree.column(col, width=160, anchor='w')
tree.grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=(100, 10))

# Definiere Tags für abwechselnde Zeilenfarben
tree.tag_configure('evenrow', background='#E8E8E8')  # Hellgrau für gerade Zeilen
tree.tag_configure('oddrow', background='white')  # Weiß für ungerade Zeilen

# Maschinen-Buttons über der Tabelle
machine_buttons_frame = ttk.Frame(content)
machine_buttons_frame.grid(row=0, column=0, sticky="nw", padx=10, pady=10)
for machine in machines:
    btn = ttk.Button(machine_buttons_frame, text=machine, style=machine_button_style,
                     command=lambda m=machine: select_machine(m))
    btn.pack(side=tk.LEFT, padx=5, pady=5)


def get_file_timestamp(file_path):
    """Gibt den Zeitstempel der letzten Änderung der Datei zurück."""
    try:
        return os.path.getmtime(file_path)
    except OSError:
        return None


# Globale Variable, um den Zeitstempel der Datei zu speichern
last_file_timestamp = None


def check_file_change():
    global last_file_timestamp
    current_timestamp = get_file_timestamp(file_path_var.get())

    if current_timestamp is None:
        messagebox.showerror("Fehler", "Die Datei konnte nicht gefunden werden.")
        return False

    if last_file_timestamp is not None and current_timestamp != last_file_timestamp:
        load_excel_data()  # Funktion zum Laden der Daten
        last_file_timestamp = current_timestamp
        messagebox.showinfo("Aktualisierung", "Die Daten wurden aktualisiert, da die Datei extern geändert wurde.")
        return False

    return True


def try_save_excel(file_path):
    global last_file_timestamp, df
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['Fräsen']  # Ersetze 'Fräsen' durch den tatsächlichen Namen Ihres Arbeitsblatts, wenn nötig

        # Schutz deaktivieren
        if sheet.protection.sheet:
            sheet.protection.disable()

        # Durchlaufen der Zeilen im DataFrame und Aktualisierung der entsprechenden Zellen im Excel-Blatt
        for index, row in df.iterrows():
            # Excel-Zeilen beginnen bei 1 und Header-Zeile wird hinzugezählt, daher `index + 2`
            # Spalte J - Werkzeug
            sheet.cell(row=index + 2, column=10).value = row[
                9]  # Nehmen Sie an, dass die 10. Spalte im DataFrame dem Werkzeug entspricht
            # Spalte L - Bemerkung
            sheet.cell(row=index + 2, column=12).value = row[
                11]  # Nehmen Sie an, dass die 12. Spalte im DataFrame der Bemerkung entspricht
            # Spalte M - Auftragsstatus
            sheet.cell(row=index + 2, column=13).value = row[
                12]  # Nehmen Sie an, dass die 13. Spalte im DataFrame dem Auftragsstatus entspricht

        # Schutz wieder aktivieren
        if sheet.protection.sheet:
            sheet.protection.enable()

        workbook.save(file_path)

        # Aktualisiere den Zeitstempel der Datei, nachdem die Änderungen erfolgreich gespeichert wurden
        last_file_timestamp = get_file_timestamp(file_path)
        return True
    except PermissionError:
        messagebox.showerror("Fehler",
                             "Derzeit kann nicht gespeichert werden. Bitte warten Sie einen Moment und versuchen Sie "
                             "es später erneut.")
        return False
    except Exception as e:
        messagebox.showerror("Unbekannter Fehler", f"Ein unbekannter Fehler ist aufgetreten: {e}")
        return False


def on_focus_in(event):
    subprocess.Popen("osk", shell=True)


# def on_focus_out(event):
# focus_label.focus_set()

def remove_focus():
    focus_button.focus()  # Setzen Sie den Fokus auf den unsichtbaren Button


# Neuer Stil für den ausgewählten Button
selected_machine_button_style = 'Selected.Machine.TButton'
style.configure(selected_machine_button_style, font=('Arial', 20, 'bold'), padding=(35, 10))
style.map(selected_machine_button_style,
          background=[('active', 'red'), ('!active', 'red')],
          foreground=[('active', 'red'), ('!active', 'red')])

# Globale Variable, um den zuletzt ausgewählten Button zu speichern
last_selected_button = None


def select_machine(machine_name):
    global last_selected_machine, last_selected_button

    # Überprüfe, ob eine Excel-Datei geladen wurde
    if not file_path_var.get():
        messagebox.showerror("Fehler", "Keine Excel-Datei geladen. Bitte laden Sie zuerst eine Datei.")
        return

    # Setze den Stil des zuletzt ausgewählten Buttons zurück
    if last_selected_button:
        last_selected_button.config(style=machine_button_style)

    last_selected_machine = machine_name
    print(f"Maschine ausgewählt: {last_selected_machine}")  # Zum Testen
    show_data(machine_name)

    # Ändere den Stil des ausgewählten Buttons
    for btn in machine_buttons_frame.winfo_children():
        if btn.cget("text") == machine_name:
            btn.config(style=selected_machine_button_style)
            last_selected_button = btn
            break


def ask_password():
    def on_submit():
        if password_entry.get() == "123":  # Setzen Sie Ihr Passwort hier
            focus_button.focus_set()
            select_excel_file()
            password_window.destroy()
        else:
            show_custom_warning("Fehler", "Falsches Passwort")
            password_window.attributes('-topmost', True)

    def show_custom_warning(title, message):
        warning_window = tk.Toplevel(password_window)
        warning_window.title(title)
        warning_window.geometry("300x150+10+10")  # Größe und Position anpassen
        warning_window.attributes('-topmost', True)

        tk.Label(warning_window, text=message, font=("Arial", 14)).pack(padx=10, pady=10)

        ok_button = tk.Button(warning_window, text="OK", command=warning_window.destroy)
        ok_button.pack(pady=10)
        ok_button.config(font=("Arial", 14), height=2, width=10)  # Größe und Schriftart des Buttons anpassen

    password_window = tk.Toplevel(root)
    password_window.title("Passwort eingeben")
    password_window.geometry("400x250+10+10")  # Größe und Position des Passwortfensters anpassen
    password_window.attributes('-topmost', True)

    password_label = ttk.Label(password_window, text="Passwort:")
    password_label.pack(padx=10, pady=10)

    password_entry = ttk.Entry(password_window, show="*")
    password_entry.pack(padx=10, pady=10)
    password_entry.bind("<FocusIn>", on_focus_in)  # Tastatur öffnen bei Fokussierung
    # password_entry.bind("<FocusOut>", on_focus_out)  # Tastatur schließen und Fokus entfernen

    submit_button = ttk.Button(password_window, text="Bestätigen", command=on_submit)
    submit_button.pack(padx=10, pady=10)


# Funktion zum Auswählen der Excel-Datei
def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.xlsm")])
    if file_path:
        file_path_var.set(file_path)
        load_excel_data()


# Laden und Anzeigen der Daten aus der ausgewählten Excel-Datei
def load_excel_data():
    if file_path_var.get():
        try:
            workbook = openpyxl.load_workbook(file_path_var.get(), data_only=True)
            sheet = workbook.active  # oder workbook['Blattname'], wenn Sie ein spezifisches Blatt bearbeiten möchten
            global df
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            df = pd.DataFrame(
                data)  # Konvertieren Sie die Daten in einen DataFrame, wenn Sie die `pandas` Funktionalität nutzen
            # möchten

            # Setze 'last_selected_machine' global, um auf den aktuellen Wert zuzugreifen
            global last_selected_machine

            # Zeige Daten basierend auf der zuletzt ausgewählten Maschine an
            if last_selected_machine:
                show_data(last_selected_machine)
            else:
                show_data(None)  # Zeige alle Daten, wenn keine Maschine ausgewählt war

        except Exception as e:
            messagebox.showerror("Fehler beim Laden der Datei", str(e))


# Funktion, um die Bildschirmtastatur zu öffnen
def open_keyboard():
    # subprocess.Popen("osk", shell=True)
    pass


# Funktion zum Finden der Kopfzeile und der Daten einer bestimmten Maschine
def get_machine_data(machine_name):
    for index, row in df.iterrows():
        if row[1] == machine_name:
            headers = row[2:]  # Kopfzeilen ab der dritten Spalte
            machine_start_index = index + 1  # Daten beginnen in der nächsten Zeile
            break
    else:
        print(f"Maschine '{machine_name}' nicht gefunden.")
        return pd.DataFrame()

    # Suche nach dem Ende des Maschinenblocks
    for end_index in range(machine_start_index, len(df)):
        # Überprüfen, ob die gesamte Zeile leer ist
        if df.iloc[end_index, 2:].isna().all():
            break
    else:
        end_index = len(df)

    # Extrahiere die Daten für die Maschine
    machine_data = df.iloc[machine_start_index:end_index, 2:]
    machine_data.columns = headers.fillna('')

    # Ersetze NaN-Werte durch leere Strings
    machine_data = machine_data.fillna('')

    return machine_data.reset_index(drop=True)


# Funktion zum Anzeigen der Daten für die ausgewählte Maschine
def show_data(machine_name):
    global last_selected_machine
    last_selected_machine = machine_name
    tree.delete(*tree.get_children())
    if machine_name:
        machine_data = get_machine_data(machine_name)
    else:
        machine_data = df

    for index, row in machine_data.iterrows():
        status = row.get("Status", "")
        auftragsstatus = row.get("Auftragsstatus", "")

        # Überprüfe, ob die Spalte für den Artikel einen Wert enthält
        artikel = row.get("Artikel", "")
        if artikel == "":
            continue  # Überspringe diese Zeile, da kein Artikel vorhanden ist

        # Bestimme Tags basierend auf dem Auftragsstatus
        tag = ''
        if auftragsstatus == 'Erledigt':
            tag = 'erledigt'
        elif auftragsstatus == 'Problem':
            tag = 'problem'
        else:
            # Anwenden von Zebra-Streifen für Zeilen ohne speziellen Status
            tag = 'evenrow' if index % 2 == 0 else 'oddrow'

        tree.insert('', tk.END, values=(
            artikel,
            row.get("Spannung", ""),
            row.get("Stückzahl", ""),
            row.get("Rohling ", ""),
            status,
            row.get("Wkz", ""),
            row.get("Bemerkung", ""),
            auftragsstatus
        ), tags=(tag,))

    # Konfiguriere die Tags
    tree.tag_configure('evenrow', background='#E8E8E8')
    tree.tag_configure('oddrow', background='white')
    tree.tag_configure('erledigt', background='green')
    tree.tag_configure('problem', background='red')


# Funktion zum Speichern der Bemerkung und Aktualisierung der Anzeige
def save_bemerkung():
    if not check_file_change():
        return  # Abbrechen, wenn die Datei geändert wurde

    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Warnung", "Bitte wählen Sie einen Artikel aus.")
        return

    bemerkung_text = bemerkung_entry.get("1.0", tk.END).strip()
    if not bemerkung_text:
        messagebox.showwarning("Warnung", "Keine Bemerkung eingegeben.")
        return

    selected_item = selected_items[0]
    selected_article_id = tree.item(selected_item, 'values')[0]

    for machine_name in ['DMU 80', 'DMU 90', 'DUO 1', 'DUO 2']:
        machine_data = get_machine_data(machine_name)
        df_index_list = machine_data.index[machine_data['Artikel'].astype(str) == str(selected_article_id)].tolist()

        if df_index_list:
            df_index = df_index_list[0] + df.index[df.iloc[:, 1] == machine_name].tolist()[0] + 1
            bemerkung_index = 11  # Index für "Bemerkung" in der DataFrame

            # Bemerkung aktualisieren
            df.iloc[df_index, bemerkung_index] = bemerkung_text

            # Prüfe, ob die Excel-Datei gespeichert werden kann
            if not try_save_excel(file_path_var.get()):
                return  # Beendet die Funktion, wenn die Datei nicht gespeichert werden konnte

            # Aktualisiere die Anzeige der Zeile im Treeview
            updated_values = list(tree.item(selected_item, 'values'))
            updated_values[6] = bemerkung_text  # Index der Bemerkung in der Treeview
            tree.item(selected_item, values=updated_values)

            # Leere die Bemerkungsbox
            bemerkung_entry.delete("1.0", tk.END)

            messagebox.showinfo("Erfolg", "Bemerkung erfolgreich gespeichert.")
            break
        else:
            print("Artikel nicht in den Daten von", machine_name, "gefunden.")

    else:
        messagebox.showerror("Fehler", "Artikel wurde nicht im DataFrame gefunden.")


# Funktion zum Aktualisieren des Status und der Bemerkungen
def update_status_and_bemerkung(auftragsstatus):
    if not check_file_change():
        return  # Abbrechen, wenn die Datei geändert wurde

    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Auswahl", "Bitte wählen Sie einen Artikel aus.")
        return

    selected_item = selected_items[0]
    selected_article_id = tree.item(selected_item, 'values')[0]

    for machine_name in ['DMU 80', 'DMU 90', 'DUO 1', 'DUO 2']:
        machine_data = get_machine_data(machine_name)

        df_index_list = machine_data.index[machine_data['Artikel'].astype(str) == str(selected_article_id)].tolist()

        if df_index_list:
            # Index in der DataFrame
            df_index = df_index_list[0] + df.index[df.iloc[:, 1] == machine_name].tolist()[0] + 1

            # Auftragsstatus Index in der DataFrame (18 ist der Index in Excel, also 17 in Python)
            auftragsstatus_df_index = 12

            if auftragsstatus == "Zurücksetzen":
                # Setze nur den Auftragsstatus in der DataFrame auf einen leeren String
                df.iloc[df_index, auftragsstatus_df_index] = ''
                updated_values = list(tree.item(selected_item, 'values'))
                updated_values[7] = ''  # Auftragsstatus-Index in der Treeview-Liste
            else:
                # Setze den Auftragsstatus und die Bemerkung in der DataFrame
                df.iloc[df_index, auftragsstatus_df_index] = auftragsstatus
                bemerkung = bemerkung_entry.get("1.0", tk.END).strip()
                df.iloc[df_index, 11] = bemerkung  # Bemerkungs-Index in der DataFrame
                updated_values = list(tree.item(selected_item, 'values'))
                updated_values[7] = auftragsstatus  # Auftragsstatus-Index in der Treeview-Liste

            # Aktualisiere die Anzeige im Treeview
            tree.item(selected_item, values=updated_values)

            # Prüfe, ob die Excel-Datei gespeichert werden kann
            if not try_save_excel(file_path_var.get()):
                return  # Beendet die Funktion, wenn die Datei nicht gespeichert werden konnte

            # Aktualisiere den Tag im Treeview
            tag = 'erledigt' if auftragsstatus == 'Erledigt' else 'problem' if auftragsstatus == 'Problem' else ''
            tree.item(selected_item, tags=(tag,))
            messagebox.showinfo("Erfolg", "Auftragsstatus aktualisiert.")
            break
        else:
            print(f"Artikel {selected_article_id} nicht in den Daten von {machine_name} gefunden.")
    else:
        messagebox.showerror("Fehler", "Artikel wurde nicht im DataFrame gefunden.")


# Funktion zum Löschen der Bemerkung für den ausgewählten Artikel
def clear_bemerkung():
    if not check_file_change():
        return  # Abbrechen, wenn die Datei geändert wurde

    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Warnung", "Bitte wählen Sie einen Artikel aus.")
        return

    selected_item = selected_items[0]
    selected_article_id = tree.item(selected_item, 'values')[0]

    for machine_name in ['DMU 80', 'DMU 90', 'DUO 1', 'DUO 2']:
        machine_data = get_machine_data(machine_name)
        df_index_list = machine_data.index[machine_data['Artikel'].astype(str) == str(selected_article_id)].tolist()

        if df_index_list:
            df_index = df_index_list[0] + df.index[df.iloc[:, 1] == machine_name].tolist()[0] + 1
            df.iloc[df_index, 11] = ''  # Setze die Bemerkung in der DataFrame auf einen leeren String

            # Aktualisiere die Anzeige der Zeile im Treeview
            tree.item(selected_item, values=(tree.item(selected_item, 'values')[0:6] + ('',)))

            # Prüfe, ob die Excel-Datei gespeichert werden kann
            if not try_save_excel(file_path_var.get()):
                return  # Beendet die Funktion, wenn die Datei nicht gespeichert werden konnte

            messagebox.showinfo("Erfolg", "Bemerkung erfolgreich gelöscht.")
            break


# Stil für Sidebar-Buttons anpassen
style.configure('Sidebar.TButton', font=('Arial', 12), padding=(15, 10))

# Laden der Icons
icons = {name: ImageTk.PhotoImage(Image.open(icon_paths[name]).resize((30, 30))) for name in icon_paths}

# Erstellen der Sidebar-Buttons in einer Schleife für "Erledigt" und "Problem"
status_buttons_frame = ttk.Frame(sidebar)  # Neues Frame für die Status-Buttons
status_buttons_frame.pack(padx=0, pady=0, fill='x', expand=True)  # Platzieren des Frames in der Sidebar

for status in ["Erledigt", "Problem"]:
    icon = icons[status]
    button = ttk.Button(status_buttons_frame, text=status, image=icon, style='Sidebar.TButton', compound="top",
                        command=lambda s=status: update_status_and_bemerkung(s))
    button.image = icon  # Referenz behalten, um Garbage Collection zu verhindern
    button.pack(side=tk.LEFT, padx=0, pady=0, expand=True)  # Buttons nebeneinander anordnen

# Button "Zurücksetzen" separat in der Sidebar hinzufügen
zuruecksetzen_icon = icons["Zurücksetzen"]
zuruecksetzen_button = ttk.Button(sidebar, text="Zurücksetzen", image=zuruecksetzen_icon, style='Sidebar.TButton',
                                  compound="top", command=lambda: update_status_and_bemerkung("Zurücksetzen"))
zuruecksetzen_button.image = zuruecksetzen_icon
zuruecksetzen_button.pack(padx=0, pady=0, fill='x')  # Button unterhalb der anderen Status-Buttons


def toggle_wkz():
    if not check_file_change():
        return  # Abbrechen, wenn die Datei geändert wurde

    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Auswahl", "Bitte wählen Sie einen Artikel aus.")
        return

    selected_item = selected_items[0]
    selected_article_id = tree.item(selected_item, 'values')[0]

    for machine_name in ['DMU 80', 'DMU 90', 'DUO 1', 'DUO 2']:
        machine_data = get_machine_data(machine_name)
        df_index_list = machine_data.index[machine_data['Artikel'].astype(str) == str(selected_article_id)].tolist()

        if df_index_list:
            df_index = df_index_list[0] + df.index[df.iloc[:, 1] == machine_name].tolist()[0] + 1
            wkz_excel_index = 9  # Index für "Wkz" in der Excel-Datei

            # Bestimme den neuen Wkz-Status
            current_wkz = df.iloc[df_index, wkz_excel_index]
            new_wkz = '' if current_wkz == 'X' else 'X'

            # Aktualisiere den DataFrame, aber noch nicht die Anzeige im Treeview
            df.iloc[df_index, wkz_excel_index] = new_wkz

            # Versuche, die Excel-Datei zu speichern
            if try_save_excel(file_path_var.get()):
                # Aktualisiere die Anzeige der Zeile im Treeview, wenn das Speichern erfolgreich war
                wkz_treeview_index = tree["columns"].index("Wkz")  # Index für "Wkz" im Treeview
                updated_values = list(tree.item(selected_item, 'values'))
                updated_values[wkz_treeview_index] = new_wkz
                tree.item(selected_item, values=updated_values)

                messagebox.showinfo("Erfolg", "Wkz erfolgreich aktualisiert.")
            else:
                # Wenn das Speichern fehlschlägt, rolle die Änderungen im DataFrame zurück und zeige eine
                # Fehlermeldung an
                df.iloc[df_index, wkz_excel_index] = current_wkz
                # Fehlermeldung wird direkt in try_save_excel angezeigt, keine Notwendigkeit, sie hier noch einmal zu
                # zeigen
            break
    else:
        messagebox.showerror("Fehler", "Artikel wurde nicht im DataFrame gefunden.")


# Button zum Setzen/Entfernen des Kreuzes in "Wkz"
toggle_wkz_button = ttk.Button(sidebar, text="Wkz Ein/Aus", style=sidebar_button_style, command=toggle_wkz)
toggle_wkz_button.pack(padx=0, pady=5, fill='x')

bemerkung_entry_width = 20  # Setzen Sie hier die gewünschte Breite
bemerkung_entry_height = 9  # Setzen Sie hier die gewünschte Höhe

# Eingabefeld für Bemerkungen mit Bildschirmtastatur-Unterstützung
bemerkung_entry = tk.Text(sidebar, height=bemerkung_entry_height, width=bemerkung_entry_width, font=("Arial", 18))
bemerkung_entry.bind("<FocusIn>", on_focus_in)
bemerkung_entry.bind("<FocusOut>", lambda e: remove_focus())  # Füge diesen Bind hinzu
bemerkung_entry.pack(padx=0, pady=10, fill='x')

# Buttons auf der rechten Seite
# Stellen Sie sicher, dass die Buttons groß genug sind für Touch-Interaktion
button_height = 3  # Höhe der Buttons anpassen
button_width = 15  # Breite der Buttons anpassen

# Bemerkung speichern Button
bemerkung_confirm_button = ttk.Button(sidebar, text="Bemerkung speichern", style=sidebar_button_style,
                                      command=save_bemerkung)
bemerkung_confirm_button.pack(padx=0, pady=5, fill='x')

# Button zum Löschen der Bemerkung
clear_bemerkung_button = ttk.Button(sidebar, text="Bemerkung löschen", style=sidebar_button_style,
                                    command=clear_bemerkung)
clear_bemerkung_button.pack(padx=0, pady=5, fill='x')

# Label für die Anzeige der letzten Aktualisierung, platziert am unteren Rand des Fensters
last_updated_label = ttk.Label(root, text="Letzte Aktualisierung: -", font=("Arial", 8), borderwidth=0, relief="flat")
last_updated_label.place(relx=1.0, rely=1.0, anchor="se", height=30)


def reload_excel_data():
    global last_selected_machine
    try:
        # Speichere den Zustand der zuletzt ausgewählten Maschine
        saved_machine = last_selected_machine

        # Lade die Daten neu
        load_excel_data()

        # Stelle den gespeicherten Zustand wieder her
        if saved_machine:
            show_data(saved_machine)
        else:
            show_data(None)  # Zeige alle Daten, wenn keine Maschine ausgewählt war

        # Aktualisiere die Anzeige der letzten Aktualisierung
        current_time = time.strftime("%H:%M:%S")
        last_updated_label.config(text=f"Letzte Aktualisierung: {current_time}")
        print("Excel-Datei neu geladen um:", current_time)
    except Exception as e:
        print("Fehler beim Neuladen der Excel-Datei:", e)


def auto_reload_excel():
    # Rufe die Funktion nach einer festgelegten Zeit erneut auf
    root.after(300000, auto_reload_excel)  # 5 Minuten Intervall
    reload_excel_data()


# Starte den automatischen Neuladeprozess nach dem Initialisieren der GUI
root.after(300000, auto_reload_excel)  # Starte nach 5 Minuten


def open_keyboard():
    subprocess.Popen("osk", shell=True)


def ask_exit_password():
    def on_submit():
        if exit_password_entry.get() == "123":  # Setzen Sie Ihr Schließpasswort hier
            root.destroy()
        else:
            messagebox.showerror("Falsches Passwort",
                                 "Das eingegebene Passwort ist falsch. Bitte versuchen Sie es erneut.",
                                 parent=exit_password_window)
            exit_password_entry.delete(0, tk.END)
            open_keyboard()  # Öffnet erneut die Tastatur, wenn das Passwort falsch war

    exit_password_window = tk.Toplevel()
    exit_password_window.title("Passwort zum Beenden")
    exit_password_window.geometry("300x250")  # Größe anpassen
    exit_password_window.attributes('-topmost', True)

    tk.Label(exit_password_window, text="Passwort zum Beenden:").pack(padx=10, pady=10)
    exit_password_entry = ttk.Entry(exit_password_window, show="*")
    exit_password_entry.pack(padx=10, pady=10)
    exit_password_entry.bind("<FocusIn>", lambda event: open_keyboard())
    exit_password_entry.focus()  # Setzt den Fokus auf das Eingabefeld

    submit_button = ttk.Button(exit_password_window, text="Bestätigen", command=on_submit)
    submit_button.pack(pady=10)


def on_close_request():
    ask_exit_password()


# Überschreibe die Standard-Schließfunktion des Hauptfensters
root.protocol("WM_DELETE_WINDOW", on_close_request)

# Neues Frame für die Buttons in der Sidebar
buttons_frame = ttk.Frame(sidebar)
buttons_frame.pack(side=tk.TOP, padx=0, pady=10)  # Packen Sie das Frame an der oberen Seite der Sidebar

# Excel-Datei Button
select_file_button = ttk.Button(buttons_frame, text="Excel Import", image=icons["Excel-Datei"],
                                style=sidebar_button_style, compound="top", command=ask_password)
select_file_button.image = icons["Excel-Datei"]
select_file_button.pack(side=tk.LEFT, padx=(0, 20), pady=(0, 30))  # Platzieren Sie den Button im Frame

# Tastatur-Button
keyboard_icon_image = ImageTk.PhotoImage(
    Image.open("tastatur.png").resize((30, 30)))  # Laden und Größenanpassung des Icons
keyboard_button = ttk.Button(buttons_frame, image=keyboard_icon_image, command=open_keyboard)
keyboard_button.image = keyboard_icon_image  # Speichern der Referenz
keyboard_button.pack(side=tk.LEFT, padx=40, pady=(5, 19))  # Platzieren Sie den Button neben dem Excel-Datei-Button


def ask_fullscreen_exit_password():
    def on_submit():
        if exit_password_entry.get() == "123":  # Ersetzen Sie 'IhrPasswort' durch Ihr tatsächliches Passwort
            root.attributes("-fullscreen", False)
            exit_password_window.destroy()
        else:
            messagebox.showerror("Falsches Passwort",
                                 "Das eingegebene Passwort ist falsch. Bitte versuchen Sie es erneut.",
                                 parent=exit_password_window)
            exit_password_entry.delete(0, tk.END)

    exit_password_window = tk.Toplevel()
    exit_password_window.title("Passwort zum Beenden des Vollbildmodus")
    exit_password_window.geometry("300x250")
    exit_password_window.attributes('-topmost', True)

    tk.Label(exit_password_window, text="Passwort:").pack(padx=10, pady=10)
    exit_password_entry = ttk.Entry(exit_password_window, show="*")
    exit_password_entry.pack(padx=10, pady=10)
    exit_password_entry.focus()

    submit_button = ttk.Button(exit_password_window, text="Bestätigen", command=on_submit)
    submit_button.pack(pady=10)


# Funktion zum Beenden des Vollbildmodus
def toggle_fullscreen(event=None):
    root.attributes("-fullscreen", False)


# Binden der Escape-Taste, um den Vollbildmodus zu beenden
root.bind("<Escape>", lambda event: ask_fullscreen_exit_password())

# Startet die Tkinter Event-Schleife
root.mainloop()
